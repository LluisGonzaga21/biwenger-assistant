"""
Genera el Excel de analisis: mercado + plantillas rivales + tu plantilla,
con ratios puntos/valor de mercado y puntos/clausula (temporada actual y,
si la API la trae, la temporada anterior).

IMPORTANTE: la API de Biwenger no es oficial ni esta documentada, asi que
los nombres de campo (ver CAMPOS_CANDIDATOS en biwenger_helpers.py) son la
mejor aproximacion posible a partir de clientes de la comunidad. Si al
ejecutar ves avisos de "campo no encontrado", ejecuta antes
01_explorar_api.py, mira el JSON real en output/debug/ y ajustamos juntos
esas listas.

Usa la misma cache local que el notebook de analisis (output/cache/*.json)
-- la primera vez descarga todo de la API (puede tardar varios minutos por
el limite de peticiones), las siguientes es practicamente instantaneo.

Uso:
    1. Copia .env.example a .env y rellenalo.
    2. (recomendado la primera vez) python 01_explorar_api.py  -> valida la conexion
    3. python 02_generar_excel.py            -> usa la cache si ya existe
       python 02_generar_excel.py --refresh  -> vuelve a descargar todo de la API
    4. El Excel se genera en output/analisis_biwenger_<fecha>.xlsx
"""

from __future__ import annotations

import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from biwenger_helpers import cargar_datos_liga, construir_dataframe

# evita que un simbolo no-ASCII en un print tumbe el script en consolas que
# no usan UTF-8 (p.ej. cmd.exe con la codepage por defecto) -- lo sustituye
# en vez de lanzar UnicodeEncodeError
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(errors="replace")

load_dotenv()

EMAIL = os.getenv("BIWENGER_EMAIL")
PASSWORD = os.getenv("BIWENGER_PASSWORD")
LEAGUE_ID = os.getenv("BIWENGER_LEAGUE_ID") or None
OWN_TEAM_ID = os.getenv("BIWENGER_OWN_TEAM_ID") or None  # override manual si el auto-detect falla

OUTPUT_DIR = Path(__file__).parent / "output"

COLUMNAS_RATIO = [
    "Ratio pts/VM (actual)", "Ratio pts/VM (anterior)",
    "Ratio pts totales/clausula (actual)", "Ratio pts totales/clausula (anterior)",
    "Ratio pts medios/clausula (actual)", "Ratio pts medios/clausula (anterior)",
]


def hoja_con_formato(writer, df: pd.DataFrame, nombre_hoja: str, columnas_ratio: list):
    if df.empty:
        df = pd.DataFrame({"Sin datos": []})
    elif "Jugador" in df.columns:
        # 'Jugador' la primera columna, para poder fijarla y que el nombre
        # se vea siempre aunque hagas scroll a la derecha
        resto = [c for c in df.columns if c != "Jugador"]
        df = df[["Jugador"] + resto]

    df.to_excel(writer, sheet_name=nombre_hoja, index=False)
    ws = writer.sheets[nombre_hoja]

    # fija la fila de cabecera y la columna 'Jugador' (columna A) a la vez
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max([len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str).tolist()[:200]])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)

    for col_name in columnas_ratio:
        if col_name not in df.columns:
            continue
        col_idx = df.columns.get_loc(col_name) + 1
        col_letter = get_column_letter(col_idx)
        rango = f"{col_letter}2:{col_letter}{len(df) + 1}"
        rule = ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        )
        ws.conditional_formatting.add(rango, rule)


def main():
    if not EMAIL or not PASSWORD:
        sys.exit(
            "Faltan BIWENGER_EMAIL / BIWENGER_PASSWORD.\n"
            "Copia .env.example a .env y rellenalo con tus datos de Biwenger."
        )

    forzar_refresco = "--refresh" in sys.argv
    try:
        data = cargar_datos_liga(
            EMAIL, PASSWORD, league_id=LEAGUE_ID, own_team_id=OWN_TEAM_ID,
            forzar_refresco=forzar_refresco,
        )
    except RuntimeError as e:
        sys.exit(str(e))

    print(f"liga: {LEAGUE_ID or '(auto)'}, tu team_id: {data['mi_team_id']}, "
          f"score_id: {data['score_id']}, saldo: {data['balance']:,}")

    df = construir_dataframe(data)

    df_mercado = df[df["Origen"] == "Mercado"]
    df_rivales = df[df["Origen"].str.startswith("Rival:")]
    df_mias = df[df["Origen"] == "Mi equipo"]

    if not df_mercado.empty and "Ratio pts/VM (actual)" in df_mercado:
        df_mercado = df_mercado.sort_values("Ratio pts/VM (actual)", ascending=False)
    if not df_rivales.empty and "Ratio pts/VM (actual)" in df_rivales:
        df_rivales = df_rivales.sort_values("Ratio pts/VM (actual)", ascending=False)
    if not df_mias.empty and "Ratio pts/VM (actual)" in df_mias:
        df_mias = df_mias.sort_values("Ratio pts/VM (actual)", ascending=False)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    fecha = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = OUTPUT_DIR / f"analisis_biwenger_{fecha}.xlsx"

    print(f"Generando Excel en {out_path}...")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        hoja_con_formato(writer, df_mercado, "Mercado", COLUMNAS_RATIO)
        hoja_con_formato(writer, df_rivales, "Rivales", COLUMNAS_RATIO)
        hoja_con_formato(writer, df_mias, "Mi equipo", COLUMNAS_RATIO)

    print(f"\nListo -> {out_path}")


if __name__ == "__main__":
    main()
